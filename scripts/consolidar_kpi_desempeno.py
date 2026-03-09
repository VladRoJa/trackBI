import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[1]  # TRACK BI
IN_DIR = (BASE_DIR / "data" / "desempeno").resolve()
OUT_LONG_XLSX = (IN_DIR / "kpi_desempeno_consolidado_long.xlsx").resolve()
OUT_LONG_CSV = (IN_DIR / "kpi_desempeno_consolidado_long.csv").resolve()

DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")

def parse_date_from_filename(p: Path):
    m = DATE_RE.search(p.name)
    return m.group(1) if m else None

def find_table_in_workbook(xlsx_path: Path):
    """
    Busca en todas las hojas una fila header que contenga 'Sucursal'
    y regresa (df, sheet_name) si encuentra una tabla válida.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    for ws in wb.worksheets:
        # scan razonable (ajusta si tus hojas son gigantes)
        max_r = min(ws.max_row or 1, 200)
        max_c = min(ws.max_column or 1, 80)

        header_row = None
        header_col_idx = None

        # 1) encontrar la celda "Sucursal"
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                if str(v).strip().lower() == "sucursal":
                    header_row = r
                    header_col_idx = c
                    break
            if header_row:
                break

        if not header_row:
            continue  # probar otra hoja

        # 2) construir headers desde esa fila (de col header_col_idx hasta max_c)
        headers = []
        for c in range(header_col_idx, max_c + 1):
            v = ws.cell(header_row, c).value
            v = "" if v is None else str(v).strip()
            headers.append(v)

        # corta headers vacíos al final
        while headers and headers[-1] == "":
            headers.pop()

        if not headers or headers[0].lower() != "sucursal":
            continue

        # 3) leer filas hacia abajo hasta que esté “vacío” varias veces
        rows = []
        empty_streak = 0
        for r in range(header_row + 1, (ws.max_row or header_row) + 1):
            row_vals = []
            for c in range(header_col_idx, header_col_idx + len(headers)):
                row_vals.append(ws.cell(r, c).value)

            # criterio de corte: fila totalmente vacía
            if all(v is None or str(v).strip() == "" for v in row_vals):
                empty_streak += 1
                if empty_streak >= 5:  # 5 filas vacías seguidas => fin tabla
                    break
                continue
            empty_streak = 0
            rows.append(row_vals)

        if not rows:
            continue

        df = pd.DataFrame(rows, columns=headers)

        # limpieza: quitar columnas vacías/Unnamed
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how="all")
        df = df.loc[:, ~df.columns.str.contains(r"^Unnamed", case=False, na=False)]
        df["Sucursal"] = df["Sucursal"].astype(str).str.strip()

        # validar que “Sucursal” tiene datos reales
        if df["Sucursal"].replace({"None": "", "nan": ""}).str.strip().eq("").all():
            continue

        return df, ws.title

    return None, None

def main():
    if not IN_DIR.exists():
        raise SystemExit(f"No existe la carpeta: {IN_DIR}")

    files = sorted(IN_DIR.glob("*.xlsx"))
    if not files:
        raise SystemExit(f"No encontré .xlsx en: {IN_DIR}")

    out_rows = []

    for f in files:
        fecha = parse_date_from_filename(f)

        df, sheet = find_table_in_workbook(f)
        if df is None:
            print(f"[SKIP] {f.name}: no encontré tabla con columna 'Sucursal' en ninguna hoja.")
            continue

        # value_vars = todas las métricas menos Sucursal
        id_vars = ["Sucursal"]
        value_vars = [c for c in df.columns if c not in id_vars]

        long_df = df.melt(
            id_vars=id_vars,
            value_vars=value_vars,
            var_name="kpi",
            value_name="valor",
        )

        long_df.insert(0, "fecha_corte", fecha)
        long_df.insert(1, "hoja", sheet)

        # a numérico
        long_df["valor"] = pd.to_numeric(long_df["valor"], errors="coerce")
        long_df["kpi"] = long_df["kpi"].astype(str).str.strip()

        long_df = long_df.dropna(subset=["valor"], how="all")

        out_rows.append(long_df)
        print(f"[OK] {f.name} (sheet='{sheet}') -> filas long: {len(long_df)} | fecha={fecha}")

    if not out_rows:
        raise SystemExit("No se consolidó nada (todos los archivos fueron SKIP).")

    final_df = pd.concat(out_rows, ignore_index=True)

    # orden
    final_df = final_df.sort_values(["fecha_corte", "Sucursal", "kpi"])

    final_df.to_excel(OUT_LONG_XLSX, index=False)
    final_df.to_csv(OUT_LONG_CSV, index=False, encoding="utf-8-sig")

    print("\n✅ Consolidado listo:")
    print(f"   - {OUT_LONG_XLSX}")
    print(f"   - {OUT_LONG_CSV}")

if __name__ == "__main__":
    main()